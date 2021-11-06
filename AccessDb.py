# This script is for accessing the Bariatric CPAP Db excel spreadsheet
from openpyxl import load_workbook
import pandas as pd
from RecordsDb import *
from datetime import datetime

def load_sheet(location, sheet_name):
    """loads the sheet excel doc and returns the compliance sheet"""
    db = load_workbook(location, read_only=True, data_only=True)
    return db[sheet_name]


def create_patient_db(compliance_sheet, outcomes_sheet):
    """Excel spreadsheet columns of interest"""
    Patients = RecordsDb()
    Patients = process_Outcomes_Db(Patients, outcomes_sheet)
    Patients = process_Adherence_db(Patients, compliance_sheet)
    return Patients


def process_Adherence_db(Patients, compliance_sheet):
    """takes patient db and fills it with info based on compliance sheet"""
    MRN_Column = 0  # "A"
    Diag_AHI_Column = 16  # "Q"
    Diag_AHI_Date_Column = 18 # "S"
    Percent_Days_4_Hours_Column = 5  # "F"
    Num_Days_Reported_Column = 6  # "G"
    Download_Date_Column = 1  # "B"
    i = 1

    print("Processing Adherence")

    compliance_iter = compliance_sheet.iter_rows()
    next(compliance_iter)  # skip first row
    for patient_row in compliance_iter:
        try:
            # Uses real MRN to cross-index the two databases
            row_mrn = int(patient_row[MRN_Column].value)
        except (ValueError, TypeError):
            row_mrn = None
        try:
            row_ahi = float(patient_row[Diag_AHI_Column].value)
        except (ValueError, TypeError):
            row_ahi = None
        try:
            row_ahi_date = patient_row[Diag_AHI_Date_Column].value
        except (ValueError, TypeError):
            row_ahi_date = None
        try:
            percent_days_4_hours = float(patient_row[Percent_Days_4_Hours_Column].value)
        except (ValueError, TypeError):
            percent_days_4_hours = None
        try:
            num_days_reported = int(patient_row[Num_Days_Reported_Column].value)
        except (ValueError, TypeError):
            num_days_reported = None
        try:
            download_date = patient_row[Download_Date_Column].value
        except (ValueError, TypeError):
            download_date = None

        row_compliance = complianceReport(download_date, num_days_reported,
            percent_days_4_hours)

        #print("Processing adherence record number " + str(i-1))
        Patient = Patients.findPatient(row_mrn)
        if Patient is None:
            # Patient is not already in list
            print("Not already on list, adding")
            Patient = PatientRecord(row_mrn)
            Patients.addPatient(Patient)
        Patient.setAHI(row_ahi, row_ahi_date)
        Patient.addComplianceRecord(row_compliance)
        i = i+1
    return Patients


def process_Outcomes_Db(Patients, outcomes_sheet):
    """takes patient db and fills it based on outcomes excel spreadsheet"""
    # TODO: add Sleep study flag? there should be 133 and 119 with use documented. - why different from adherence-based?

    print("Processing Outcomes")
    outcomes_iter = outcomes_sheet.iter_rows()
    title_row = next(outcomes_iter)  # skip 1st/labels for the iteration, save title row to generate label numbers.
    # Generate the column numbers that correspond to each label so that each row can be accessed.
    for i in range(len(title_row)):
        if title_row[i].value == "PAT_ID":
            MRN_Column = i
        if title_row[i].value == "BARI_SURGERY_DATE":
            Bari_Surg_Date_Column = i
        if title_row[i].value == "HEIGHT_CM_DOS":
            Height_DOS_Column = i
        if title_row[i].value == "WEIGHT_KG_DOS":
            Weight_DOS_Column = i
        if title_row[i].value == "WEIGHT_KG_2_MONTH":
            Weight_2mo_Column = i
        if title_row[i].value == "WEIGHT_KG_4_MONTH":
            Weight_4mo_Column = i
        if title_row[i].value == "WEIGHT_KG_6_MONTH":
            Weight_6mo_Column = i
        if title_row[i].value == "WEIGHT_KG_1_YEAR":
            Weight_1y_Column = i
        if title_row[i].value == "WEIGHT_KG_2_YEAR":
            Weight_2y_Column = i
        if title_row[i].value == "WEIGHT_KG_3_YEAR":
            Weight_3y_Column = i
        if title_row[i].value == "WEIGHT_KG_4_YEAR":
            Weight_4y_Column = i
        if title_row[i].value == "WEIGHT_KG_5_YEAR":
            Weight_5y_Column = i
        if title_row[i].value == "Pre-op date":
            pre_op_date_Column = i
            # print("pre-op date column " + str(i))
        if title_row[i].value == "HCO3 pre-op":
            HCO3_pre_op_Column = i
            # print("pre-op hco3 column " + str(i))
        if title_row[i].value == "sCr pre-op":
            sCr_pre_op_Column = i
            # print("pre-op scr column " + str(i))
        if title_row[i].value == "HCO3 before":   # note, this is actually day 1 - 14 post-op
            HCO3_DOS_Column = i
        if title_row[i].value == "HCO3 2 months":
            HCO3_2mo_Column = i
        if title_row[i].value == "HCO3 4 months":
            HCO3_4mo_Column = i
        if title_row[i].value == "HCO3 6 months":
            HCO3_6mo_Column = i
        if title_row[i].value == "HCO3 1 year":
            HCO3_1y_Column = i
        if title_row[i].value == "HCO3 2 year":
            HCO3_2y_Column = i
        if title_row[i].value == "HCO3 3 year":
            HCO3_3y_Column = i
        if title_row[i].value == "HCO3 4 year":
            HCO3_4y_Column = i
        if title_row[i].value == "HCO3 5 year":
            HCO3_5y_Column = i
        if title_row[i].value == "Sex":
            sex_column = i
        if title_row[i].value == "DOB":
            dob_column = i
        if title_row[i].value == "CCI_SCORE":
            cci_column = i
        if title_row[i].value == "Exlusion Loop":
            loop_exclusion_column = i
        if title_row[i].value == 'Exclusion Opiate':
            opiate_exclusion_column = i
        if title_row[i].value == 'Exclusion CA':
            ca_exclusion_column = i
        if title_row[i].value == 'Exclusion Other':
            other_exclusion_column = i
    try:
        weightOrder = [Weight_DOS_Column, Weight_2mo_Column, Weight_4mo_Column,
            Weight_6mo_Column, Weight_1y_Column, Weight_2y_Column, Weight_3y_Column,
            Weight_4y_Column, Weight_5y_Column]
        HCO3Order = [HCO3_pre_op_Column, HCO3_2mo_Column, HCO3_4mo_Column,
            HCO3_6mo_Column, HCO3_1y_Column, HCO3_2y_Column, HCO3_3y_Column,
            HCO3_4y_Column, HCO3_5y_Column] # Note: HCO3_DOS_Column removed
    except(UnboundLocalError) as e:
        print("Error: No column labels found for one or more of the needed columns")
        print(e)
        quit()

    for patient_row in outcomes_iter:
        # For each row that has an MRN entry...
        #print("Processing chart #" + str(i))
        try:
            row_mrn = int(patient_row[MRN_Column].value)
        except(ValueError, TypeError):
            row_mrn = None
            print("Error!")

        row_bari_dos = patient_row[Bari_Surg_Date_Column].value

        # Sex
        try:
            row_sex = str(patient_row[sex_column].value)
        except(ValueError, TypeError):
            row_sex = None

        # DOB
        try:
            row_dob = patient_row[dob_column].value
        except(ValueError, TypeError):
            row_dob = None

        # CCI
        try:
            row_cci = int(patient_row[cci_column].value)
        except(ValueError, TypeError):
            row_cci = None

        # HEIGHT
        try:
            row_height_dos = float(patient_row[Height_DOS_Column].value)
        except(ValueError, TypeError):
            row_height_dos = None

        # WEIGHT
        row_weight = list()  # list of weights; order of appending is chronological
        for column in weightOrder:
            try:
                row_weight.append(float(patient_row[column].value))
            except(ValueError, TypeError):
                row_weight.append(None)

        # BICARBONATE
        row_HCO3 = list()  # list of weights; order of appending is chronological
        for column in HCO3Order:
            try:
                row_HCO3.append(float(patient_row[column].value))
            except(ValueError, TypeError):
                row_HCO3.append(None)
        #print(row_HCO3)

        # PREOP DATE AND CREATININE
        try:
            row_date_preop = patient_row[pre_op_date_Column].value
        except(ValueError, TypeError):
            row_date_preop = None
        try:
            row_creat_preop = float(patient_row[sCr_pre_op_Column].value)
        except(ValueError, TypeError):
            row_creat_preop = None

        # EXCLUSIONS
        try:
            loop_exclusion = str(patient_row[loop_exclusion_column].value)
        except(ValueError, TypeError):
            loop_exclusion = None
        try:
            opiate_exclusion = str(patient_row[opiate_exclusion_column].value)
        except(ValueError, TypeError):
            opiate_exclusion = None
        try:
            ca_exclusion = str(patient_row[ca_exclusion_column].value)
        except(ValueError, TypeError):
            ca_exclusion = None
        try:
            other_exclusion = str(patient_row[other_exclusion_column].value)
        except(ValueError, TypeError):
            other_exclusion = None

        # HCO3 AT SURGERY
        # TODO this - low priority as it doesn't seem like there is much data to be gleaned from this

        # End of line-by-line processing
        # either add to an existing patient or create a new patient with indexed information.
        # Start of patient object creation and addition to list
        Patient = Patients.findPatient(row_mrn)
        if Patient is None:
            # Patient is not already in list
            #print("Adding patient MRN: " + str(row_mrn))
            Patient = PatientRecord(row_mrn)
            Patients.addPatient(Patient)

        if row_cci is not None:
            Patient.setCCI(row_cci)
        if row_dob is not None:
            Patient.setDOB(row_dob)
        if row_sex is not None:
            Patient.setSex(row_sex)
        if row_height_dos is not None:
            Patient.setHeight(row_height_dos)
        if row_date_preop is not None:
            Patient.setPreOpDate(row_date_preop)
        else:
            print("warning: no pre-op date")
        if row_creat_preop is not None:
            Patient.setPreOpCreat(row_creat_preop)
        if len(row_weight) is not 0:
            Patient.setWeights(row_weight)
            #print(row_weight)
        else:
            print("Warning: No weight info")
        if row_bari_dos is not None:
            Patient.setBariDOS(row_bari_dos)
        if len(row_HCO3) is not 0:
            Patient.setHCO3s(row_HCO3)
            #print(row_HCO3)
        else:
            print("Warning: No HCO3 info")
        Patient.set_exclusions(loop_exclusion, opiate_exclusion, ca_exclusion, other_exclusion)
        i = i+1
    return Patients


def test_db_gen():
    db = RecordsDb()
    patient1 = PatientRecord(1243)
    patient1.setAHI(41.0, "12/19/19")
    record1 = complianceReport("12/21/19", 90, 50)
    record2 = complianceReport("12/21/18", 30, 100)
    patient1.addComplianceRecord(record1)
    patient1.addComplianceRecord(record2)
    patient1.setHeight(185.0)
    patient1.setWeights(
        [150.0, 140.0, 140.0, 140.0, 130.0, 140.0, 145.0, None, 160.0])
    # [DOS, +2mo, +4mo, +6mo, +1y, +2y, +3y, +4y, +5y]
    print(patient1.BMIDOS())
    db.addPatient(patient1)

    patient2 = PatientRecord(1244)
    patient2.setAHI(40.0, "12/19/20")
    record3 = complianceReport("12/21/19", 90, 45)
    record4 = complianceReport("12/21/18", 30, 0)
    patient2.addComplianceRecord(record3)
    patient2.addComplianceRecord(record4)
    patient2.setHeight(150.0)
    patient2.setWeights(
        [130.0, 135.0, 140.0, 140.0, 130.0, 140.0, 145.0, 155.0, None])
    db.addPatient(patient2)

    patient3 = PatientRecord(1245)
    patient3.setAHI(None, None)
    patient3.setHeight(190.0)
    patient3.setWeights([100.0, None, None, None, None, None, None, None, None])
    db.addPatient(patient3)

    print("----")
    patient1.printPtRecord()
    patient2.printPtRecord()
    print("----")
    return db


def AccessDatabase(db_loc, compliance_db_loc, compliance_data_sheet_name,
        outcome_db_loc, outcome_data_sheet_name):
    """Accesses the two databases (compliance excel doc) and compliance excel
    doc and combines them, outputting a RecordsDb object with all records
    populated
    db_loc = the root folder that contains both excel docs
    compliance_db_loc = location of the compliance excel
    compliance_data_sheet_name
    outcome_db_loc = location of the outcome excel
    outcome_data_sheet_name
    """

    # This could have been done more elegantly w/ dataframe merge + groupby?
    compliance_sheet = load_sheet(db_loc + compliance_db_loc, compliance_data_sheet_name)
    outcomes_sheet = load_sheet(db_loc + outcome_db_loc, outcome_data_sheet_name)
    return create_patient_db(compliance_sheet, outcomes_sheet)


def main():
    # 0 for testing, 1 to run
    testing_mode = 1

    if testing_mode == 0:
        database = test_db_gen()
    else:
        pass


if __name__ == '__main__':
    main()
